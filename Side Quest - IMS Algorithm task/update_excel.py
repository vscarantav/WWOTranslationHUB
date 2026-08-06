import openpyxl
import os
import zipfile
import xml.etree.ElementTree as ET
import google.generativeai as genai
import json
import time

api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    raise ValueError("GEMINI_API_KEY environment variable not set")
genai.configure(api_key=api_key)
model = genai.GenerativeModel("gemini-3.5-flash")

mapping = {
    'CSE110': 'introducao-a-programacao-portuguese-master-export.imscc',
    'CSE111': 'programacao-com-funcoes-portuguese-master-export.imscc',
    'CSE210': 'programacao-com-classes-portuguese-dev-export.imscc',
    'CSE310': 'programacao-aplicada-portuguese-master-export.imscc',
    'WDD130': 'fundamentos-da-web-portuguese-master-export.imscc',
    'WDD131': 'fundamentos-da-web-dinamica-portuguese-pilot-export.imscc',
    'WDD231': 'desenvolvimento-frontend-para-web-i-portuguese-dev-export.imscc',
    'ITM111': 'ptbr-translation-itm111-export.imscc',
    'WDD330': 'ptbr-translation-wdd330-export.imscc'
}

excel_file = 'Software Development - IMS Actions - Copy.xlsx'
wb = openpyxl.load_workbook(excel_file)

def get_ptbr_titles(imscc_file):
    with zipfile.ZipFile(imscc_file, 'r') as z:
        manifest_data = z.read('imsmanifest.xml')
    tree = ET.fromstring(manifest_data)
    ns = {'ims': 'http://www.imsglobal.org/xsd/imsccv1p1/imscp_v1p1'}
    titles = []
    for item in tree.findall('.//ims:item', ns):
        title_elem = item.find('ims:title', ns)
        if title_elem is not None and title_elem.text:
            titles.append(title_elem.text.strip())
    return titles

for sheet_name, imscc_file in mapping.items():
    print(f"Processing {sheet_name}...")
    if sheet_name not in wb.sheetnames:
        print(f"Sheet {sheet_name} not found!")
        continue
    
    ptbr_titles = get_ptbr_titles(imscc_file)
    ws = wb[sheet_name]
    
    # Get unique English titles in Column C (Index 3)
    eng_titles = set()
    for row in range(2, ws.max_row + 1):
        val = ws.cell(row=row, column=3).value
        if val:
            eng_titles.add(str(val))
    
    if not eng_titles:
        print(f"No english titles found in {sheet_name}")
        continue
        
    print(f"Found {len(eng_titles)} unique English titles.")
    
    prompt = f"""
You are an expert translator. I have a list of English assignment names from a course report, and a list of official Portuguese (PT-BR) assignment titles from the course manifest.
Your task is to map each English assignment name to its official Portuguese title. 

Rules:
1. Some English names might have extra appended statuses (like "Question: ... Answer: ..."). If so, find the core assignment name in the PT-BR list, and append the translated status to it.
2. Some generic names like "New Student Added" or "Not Logged into Canvas" won't be in the PT-BR list. Translate them directly to natural PT-BR (e.g. "Novo Estudante Adicionado", "Não entrou no Canvas").
3. Return ONLY a valid JSON dictionary where the keys are the exact English names provided, and the values are the corresponding Portuguese translated strings. DO NOT wrap the JSON in markdown blocks, just return raw JSON.

English Assignment Names:
{json.dumps(list(eng_titles))}

Official PT-BR Titles:
{json.dumps(ptbr_titles)}
"""
    
    response = model.generate_content(prompt)
    try:
        text = response.text.strip()
        if text.startswith("```json"):
            text = text[7:-3].strip()
        elif text.startswith("```"):
            text = text[3:-3].strip()
            
        trans_map = json.loads(text)
        
        # Update the sheet
        updated_count = 0
        for row in range(2, ws.max_row + 1):
            val = ws.cell(row=row, column=3).value
            if val and str(val) in trans_map:
                ws.cell(row=row, column=3).value = trans_map[str(val)]
                updated_count += 1
        print(f"Updated {updated_count} rows in {sheet_name}")
    except Exception as e:
        print(f"Error processing {sheet_name}: {e}")
        print("Raw response:", response.text)

wb.save(excel_file)
print("Finished saving Excel.")
