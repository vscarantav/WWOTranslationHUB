import openpyxl
import json
import google.generativeai as genai

import os

# Ensure API key is set in environment or .env file
api_key = os.getenv("GEMINI_API_KEY")
if not api_key:
    raise ValueError("GEMINI_API_KEY environment variable not set")
genai.configure(api_key=api_key)
model = genai.GenerativeModel("gemini-3.5-flash")

excel_file = 'Software Development - IMS Actions - Copy.xlsx'
wb = openpyxl.load_workbook(excel_file)
sheets = ['CSE110', 'CSE111', 'CSE210', 'CSE310', 'WDD130', 'WDD131', 'WDD231', 'ITM111', 'WDD330']

unique_strings = set()
for sheet_name in sheets:
    ws = wb[sheet_name]
    for row in range(2, ws.max_row + 1):
        for col in (13, 14): # M and N
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip():
                unique_strings.add(str(val).strip())

unique_list = list(unique_strings)
print(f"Translating {len(unique_list)} unique strings...")

chunk_size = 30
translation_map = {}

import time

for i in range(0, len(unique_list), chunk_size):
    chunk = unique_list[i:i + chunk_size]
    print(f"Processing chunk {i//chunk_size + 1}/{(len(unique_list) + chunk_size - 1)//chunk_size}...")
    
    prompt = f"""
    You are an expert translator. Translate the following English strings from a course instructor report into Brazilian Portuguese (PT-BR).
    The strings are reasons for outreach, instructions for instructors on how to message students in Canvas, and course-related descriptions.
    
    Return a valid JSON dictionary where the keys are the exact English strings provided, and the values are the PT-BR translations.
    DO NOT wrap the JSON in markdown blocks, just return raw JSON text starting with {{ and ending with }}. Ensure all quotes and special characters are escaped properly.
    
    English Strings:
    {json.dumps(chunk)}
    """
    
    retries = 3
    for attempt in range(retries):
        try:
            response = model.generate_content(prompt)
            text = response.text.strip()
            if text.startswith("```json"):
                text = text[7:-3].strip()
            elif text.startswith("```"):
                text = text[3:-3].strip()
                
            trans_map = json.loads(text)
            translation_map.update(trans_map)
            break # Success, break retry loop
        except Exception as e:
            print(f"Attempt {attempt + 1} failed. Error:", e)
            if attempt < retries - 1:
                print("Retrying in 5 seconds...")
                time.sleep(5)
            else:
                print(f"Failed to process chunk {i//chunk_size + 1} after {retries} retries.")


print(f"Successfully translated {len(translation_map)} strings. Updating Excel...")

for sheet_name in sheets:
    ws = wb[sheet_name]
    updated = 0
    for row in range(2, ws.max_row + 1):
        for col in (13, 14): # M and N
            val = ws.cell(row=row, column=col).value
            if val and str(val).strip():
                key = str(val).strip()
                if key in translation_map:
                    ws.cell(row=row, column=col).value = translation_map[key]
                    updated += 1
    print(f"Updated {updated} cells in {sheet_name}")

wb.save(excel_file)
print("Finished saving Excel.")
